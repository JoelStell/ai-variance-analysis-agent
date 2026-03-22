"""
Consolidated Variance Analysis AI Agent (v5)
=============================================
Usage:
    python consolidated_narrator.py dept_submissions_input.xlsx
    python consolidated_narrator.py dept_submissions_input.xlsx -o Q2_report.xlsx
"""
import sys,os,json,argparse,getpass
from datetime import datetime
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
try:
    import anthropic; HAS_API=True
except ImportError:
    HAS_API=False; print("WARNING: pip install anthropic for AI commentary")

MPCT=0.05; MABS=500000; MODEL="claude-sonnet-4-20250514"
B=Font(name='Arial',color='000000',size=10); BB=Font(name='Arial',color='000000',size=10,bold=True)
WB=Font(name='Arial',color='FFFFFF',size=11,bold=True); TF=Font(name='Arial',color='000000',size=14,bold=True)
SF=Font(name='Arial',color='FFFFFF',size=10,bold=True); NF=Font(name='Arial',color='666666',size=9,italic=True)
GF=Font(name='Arial',color='006100',size=10); RF=Font(name='Arial',color='9C0006',size=10)
OF=Font(name='Arial',color='FF6600',size=10,bold=True); IF2=Font(name='Arial',color='999999',size=10,italic=True)
CF=Font(name='Arial',color='1F4E79',size=10); DF=Font(name='Arial',color='003366',size=9)
GB=PatternFill('solid',fgColor='C6EFCE'); RB=PatternFill('solid',fgColor='FFC7CE')
OB=PatternFill('solid',fgColor='FFF2CC'); YB=PatternFill('solid',fgColor='F2F2F2')
DB2=PatternFill('solid',fgColor='4472C4'); DH=PatternFill('solid',fgColor='44546A')
DG=PatternFill('solid',fgColor='548235'); DP=PatternFill('solid',fgColor='7030A0')
OR_H=PatternFill('solid',fgColor='C55A11'); LB=PatternFill('solid',fgColor='D6E4F0')
TB=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
AF='#,##0;(#,##0);"-"'; PF='0.0%;(0.0%);"-"'; PCF='0%'

def get_key(ck=None):
    if ck: return ck
    ek=os.environ.get('ANTHROPIC_API_KEY')
    if ek: print("  Using API key from env var."); return ek
    sd=os.path.dirname(os.path.abspath(__file__))
    for fn in ['api_key.txt','API_Key.txt','API_key.txt','Api_Key.txt']:
        kf=os.path.join(sd,fn)
        if os.path.exists(kf):
            with open(kf) as f: fk=f.read().strip()
            if fk: print(f"  Using API key from {kf}"); return fk
    print("\n  No API key. 1) Enter now 2) Enter to skip 3) API_Key.txt\n")
    k=getpass.getpass("  Paste key (hidden) or Enter to skip: ").strip()
    return k if k else None

def cv(c,p,ie=False):
    v=-(c-p) if ie else c-p; return v, v/abs(p) if p!=0 else None
def im(v,p,mp=MPCT,ma=MABS):
    return abs(v)>=ma or(p is not None and abs(p)>=mp)

def parse(fp):
    wb=load_workbook(fp,data_only=True); ds={}
    for sn in wb.sheetnames:
        ws=wb[sn]; iq=sn.endswith(' Qtr PL'); iy=sn.endswith(' YTD PL'); ib=sn.endswith(' BS')
        if not(iq or iy or ib): continue
        dn=sn.rsplit(' ',2)[0] if(iq or iy) else sn.rsplit(' ',1)[0]
        if dn not in ds: ds[dn]={'qtr':[],'ytd':[],'bs':[],'type':'Unknown'}
        hr=None
        for r in range(1,20):
            v=ws.cell(r,1).value
            if v and str(v).strip()=='Account': hr=r; break
            if v and 'Type:' in str(v):
                if 'Operating' in str(v): ds[dn]['type']='Operating Segment'
                elif 'Corporate' in str(v): ds[dn]['type']='Corporate Function'
        if hr is None: continue
        cs=None
        for r in range(hr+1,ws.max_row+1):
            a=ws.cell(r,1).value
            if not a or not str(a).strip(): continue
            a=str(a).strip()
            if a.startswith('Total ') or a in('GROSS PROFIT','OPERATING INCOME'): continue
            secs=('Revenue','Cost of Revenue','Operating Expenses','Other Income / (Expense)','Assets','Liabilities')
            if a in secs: cs=a; continue
            if iq:
                try: q2=float(ws.cell(r,2).value or 0); q1=float(ws.cell(r,3).value or 0); q2y=float(ws.cell(r,4).value or 0)
                except: continue
                ds[dn]['qtr'].append({'account':a,'section':cs or'Other','q2':q2,'q1':q1,'q2y':q2y,
                    'qoq_expl':str(ws.cell(r,7).value or'').strip(),'yoy_expl':str(ws.cell(r,10).value or'').strip()})
            elif iy:
                try: y1=float(ws.cell(r,2).value or 0); y1y=float(ws.cell(r,3).value or 0)
                except: continue
                ds[dn]['ytd'].append({'account':a,'section':cs or'Other','ytd':y1,'ytdy':y1y,
                    'ytd_expl':str(ws.cell(r,6).value or'').strip()})
            elif ib:
                try: j=float(ws.cell(r,2).value or 0); d=float(ws.cell(r,3).value or 0)
                except: continue
                ds[dn]['bs'].append({'account':a,'section':cs or'Other','jun':j,'dec':d,
                    'expl':str(ws.cell(r,6).value or'').strip()})
    wb.close(); return ds

def consol(ds,dk,afs,efs):
    c={}
    for dn,dd in ds.items():
        for i in dd[dk]:
            a=i['account']
            if a not in c:
                c[a]={'section':i['section'],'depts':[]}
                for f in afs: c[a][f]=0
                for f in efs: c[a][f]=[]
            for f in afs: c[a][f]+=i.get(f,0)
            for f in efs:
                v=i.get(f,'')
                if v and v not in('Immaterial','Immaterial.','No variance.','No variance'):
                    c[a][f].append(f'[{dn}] {v}')
            c[a]['depts'].append(dn)
    so=['Revenue','Other Income / (Expense)','Cost of Revenue','Operating Expenses','Assets','Liabilities','Other']
    r=[]; s=set()
    for sec in so:
        items=[(k,v) for k,v in c.items() if v['section']==sec and k not in s]
        if items:
            r.append(('__S__',sec))
            for k,v in items: r.append((k,v)); s.add(k)
    rem=[(k,v) for k,v in c.items() if k not in s]
    if rem:
        r.append(('__S__','Other'))
        for k,v in rem: r.append((k,v))
    return r

SYS_PROMPT = """You are a senior corporate controller preparing quarterly variance analysis for a public company 10-Q.
Generating commentary for: {st}
MATERIALITY: >{mp}% or >{ma}K

Return JSON with this EXACT structure per account:
{{"account": {{"explanation": "...", "coverage_pct": 85, "followup": "", "followup_depts": []}}}}

FIELDS:
- explanation: Synthesized variance narrative. Department attribution with dollars, ordered by magnitude.
  Format: "[Dept] (+/-$X,XXXK): driver." Concise, 2-4 sentences. CFO audience.
- coverage_pct: Integer 0-100. Honest assessment of what % of the dollar variance is explained.
- followup: ONLY if there is a problem. Include which department(s) to follow up with.
  Problems: DIRECTIONAL ERROR, MATH ERROR, CONTRADICTION, INSUFFICIENT COVERAGE (<80%), COVERAGE CLAIM ERROR.
  Format: "[Department Name]: specific issue found."
- followup_depts: Array of department names needing follow-up. Empty if clean.

RULES:
1. Every driver gets +$X,XXXK. Order by magnitude. Attribute departments.
2. Be skeptical — check math and direction against the actual numbers provided.
3. Coverage_pct must reflect reality. Not everything is 100%.
4. Do NOT put coverage % in explanation text.

RESPONSE: ONLY valid JSON. No markdown. No backticks."""

def ai_gen(items,st,mp,ma):
    if not HAS_API or not os.environ.get('ANTHROPIC_API_KEY'): return {}
    cl=anthropic.Anthropic(api_key=os.environ['ANTHROPIC_API_KEY'])
    compact=[{'a':i['account'],'v':i['var'],'p':i.get('pct',''),'d':i['dept_expls'],
              'n':i.get('nums',{})} for i in items]
    try:
        r=cl.messages.create(model=MODEL,max_tokens=4000,
            system=SYS_PROMPT.format(st=st,mp=f"{mp*100:.0f}",ma=f"{ma/1000:.0f}"),
            messages=[{"role":"user","content":f"Analyze:\n{json.dumps(compact)}"}])
        raw=r.content[0].text.strip()
        if raw.startswith('```'): raw=raw.split('\n',1)[1]
        if raw.endswith('```'): raw=raw[:-3]
        return json.loads(raw.strip())
    except Exception as e:
        print(f"  AI error: {e}"); return {}

def get_ai(cq,cy,cb,mp,ma):
    cm={}
    def mk(data,f1,f2,ef):
        items=[]
        for a,d in data:
            if a=='__S__': continue
            ie=d['section'] in('Cost of Revenue','Operating Expenses')
            v,p=cv(d[f1],d[f2],ie)
            if im(v,p,mp,ma):
                items.append({'account':a,'var':int(v),'pct':f"{p:.1%}" if p else"N/A",
                    'nums':{'curr':int(d[f1]),'prior':int(d[f2])},'dept_expls':d.get(ef,[])})
        return items
    qi=mk(cq,'q2','q1','qoq_expl'); print(f"  QoQ: {len(qi)} items")
    cm['qoq']=ai_gen(qi,"P&L QoQ (Q2 vs Q1 2025)",mp,ma) if qi else {}
    yi=mk(cq,'q2','q2y','yoy_expl'); print(f"  YoY: {len(yi)} items")
    cm['yoy']=ai_gen(yi,"P&L YoY (Q2 2025 vs Q2 2024)",mp,ma) if yi else {}
    ti=mk(cy,'ytd','ytdy','ytd_expl'); print(f"  YTD: {len(ti)} items")
    cm['ytd']=ai_gen(ti,"P&L YTD (6mo ended Jun 30 2025 vs 2024)",mp,ma) if ti else {}
    bi=[]
    for a,d in cb:
        if a=='__S__': continue
        v=d['jun']-d['dec']; p=v/abs(d['dec']) if d['dec']!=0 else None
        if im(v,p,mp,ma):
            bi.append({'account':a,'var':int(v),'pct':f"{p:.1%}" if p else"N/A",
                'nums':{'jun':int(d['jun']),'dec':int(d['dec'])},'dept_expls':d.get('expl',[])})
    print(f"  BS:  {len(bi)} items")
    cm['bs']=ai_gen(bi,"Balance Sheet (Jun 30 vs Dec 31 2024)",mp,ma) if bi else {}
    return cm

def gaf(cm,bk,ac,fld,dflt=''):
    e=cm.get(bk,{}).get(ac,{})
    if isinstance(e,dict): return e.get(fld,dflt)
    if isinstance(e,str) and fld=='explanation': return e
    return dflt

def w_expl(ws,r,c,de,ai,m):
    if not m: cl=ws.cell(r,c,'Immaterial'); cl.font=IF2; cl.fill=YB
    elif ai: cl=ws.cell(r,c,ai); cl.font=CF
    elif de: cl=ws.cell(r,c,'\n'.join(de)); cl.font=DF
    else: cl=ws.cell(r,c,''); cl.font=B
    cl.border=TB; cl.alignment=Alignment(wrap_text=True)

def w_cov(ws,r,c,pct,m):
    if not m: cl=ws.cell(r,c,''); cl.font=IF2
    elif pct is not None:
        cl=ws.cell(r,c,pct/100); cl.number_format=PCF
        if pct<80: cl.font=OF; cl.fill=OB
        else: cl.font=GF; cl.fill=GB
    else: cl=ws.cell(r,c,''); cl.font=B
    cl.border=TB; cl.alignment=Alignment(horizontal='center')

def w_fu(ws,r,c,fu):
    cl=ws.cell(r,c,fu)
    if fu: cl.font=OF; cl.fill=OB
    else: cl.font=B
    cl.border=TB; cl.alignment=Alignment(wrap_text=True)

def vc(ws,r,c,v,m):
    cl=ws.cell(r,c,v); cl.number_format=AF; cl.border=TB
    if m: cl.font=GF if v>0 else RF; cl.fill=GB if v>0 else RB
    else: cl.font=B

def build(cq,cy,cb,cm,ds,op,mp,ma):
    wb=Workbook(); ts=datetime.now().strftime("%Y-%m-%d %H:%M")
    all_fu=[]

    # ==== QTR PL ====
    # New layout: Account | Q2 25 | Q1 25 | QoQ$ | QoQ% | QoQ Expl | QoQ Cov | Q2 25 | Q2 24 | YoY$ | YoY% | YoY Expl | YoY Cov | Follow-Up
    ws=wb.active; ws.title='Consolidated Qtr PL'; ws.sheet_properties.tabColor='4472C4'
    ncol=14
    ws.merge_cells(f'A1:{get_column_letter(ncol)}1'); ws.cell(1,1,'CONSOLIDATED QUARTERLY P&L — Variance Analysis').font=TF
    ws.merge_cells(f'A2:{get_column_letter(ncol)}2'); ws.cell(2,1,f'Three Months Ended June 30, 2025 | Material: >{mp*100:.0f}% or >${ma/1000:.0f}K | {ts}').font=NF
    r=4
    hdrs=['Account','Q2 2025','Q1 2025','QoQ $','QoQ %','QoQ Explanation','QoQ\nCov%',
          'Q2 2025','Q2 2024','YoY $','YoY %','YoY Explanation','YoY\nCov%','Follow-Up Flag']
    for i,h in enumerate(hdrs,1):
        c=ws.cell(r,i,h); c.font=WB; c.fill=DB2; c.border=TB
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    # Color the QoQ group header slightly different from YoY
    for i in [8,9,10,11,12,13]:
        ws.cell(r,i).fill=PatternFill('solid',fgColor='2F5496')
    ws.row_dimensions[r].height=40; r+=1

    for a,d in cq:
        if a=='__S__':
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncol)
            ws.cell(r,1,d).font=SF
            for cc in range(1,ncol+1): ws.cell(r,cc).fill=DH; ws.cell(r,cc).border=TB
            r+=1; continue
        ie=d['section'] in('Cost of Revenue','Operating Expenses')
        qv,qp=cv(d['q2'],d['q1'],ie); yv,yp=cv(d['q2'],d['q2y'],ie)
        qm=im(qv,qp,mp,ma); ym=im(yv,yp,mp,ma)

        ws.cell(r,1,a).font=B; ws.cell(r,1).border=TB
        # QoQ block: Q2, Q1, QoQ$, QoQ%, QoQ Expl, QoQ Cov
        ws.cell(r,2,d['q2']).font=B; ws.cell(r,2).number_format=AF; ws.cell(r,2).border=TB
        ws.cell(r,3,d['q1']).font=B; ws.cell(r,3).number_format=AF; ws.cell(r,3).border=TB
        vc(ws,r,4,qv,qm)
        c=ws.cell(r,5,qp if qp else 0); c.number_format=PF; c.border=TB; c.font=B
        w_expl(ws,r,6,d.get('qoq_expl',[]),gaf(cm,'qoq',a,'explanation'),qm)
        qcov=gaf(cm,'qoq',a,'coverage_pct',None)
        w_cov(ws,r,7,qcov,qm)

        # YoY block: Q2 (repeat), Q2 PY, YoY$, YoY%, YoY Expl, YoY Cov
        ws.cell(r,8,d['q2']).font=B; ws.cell(r,8).number_format=AF; ws.cell(r,8).border=TB; ws.cell(r,8).fill=LB
        ws.cell(r,9,d['q2y']).font=B; ws.cell(r,9).number_format=AF; ws.cell(r,9).border=TB
        vc(ws,r,10,yv,ym)
        c=ws.cell(r,11,yp if yp else 0); c.number_format=PF; c.border=TB; c.font=B
        w_expl(ws,r,12,d.get('yoy_expl',[]),gaf(cm,'yoy',a,'explanation'),ym)
        ycov=gaf(cm,'yoy',a,'coverage_pct',None)
        w_cov(ws,r,13,ycov,ym)

        # Follow-up combining both
        qfu=gaf(cm,'qoq',a,'followup',''); yfu=gaf(cm,'yoy',a,'followup','')
        parts=[]
        if qfu: parts.append(f'QoQ: {qfu}')
        if yfu: parts.append(f'YoY: {yfu}')
        fu='\n'.join(parts)
        w_fu(ws,r,14,fu)

        for bk,ft in[('qoq',qfu),('yoy',yfu)]:
            if ft:
                dl=gaf(cm,bk,a,'followup_depts',[])
                all_fu.append({'tab':'Qtr PL','period':bk.upper(),'account':a,'issue':ft,
                    'depts':dl or list(set(d['depts']))})
        r+=1

    ws.column_dimensions['A'].width=35; ws.column_dimensions['B'].width=14; ws.column_dimensions['C'].width=14
    ws.column_dimensions['D'].width=12; ws.column_dimensions['E'].width=9; ws.column_dimensions['F'].width=58
    ws.column_dimensions['G'].width=8; ws.column_dimensions['H'].width=14; ws.column_dimensions['I'].width=14
    ws.column_dimensions['J'].width=12; ws.column_dimensions['K'].width=9; ws.column_dimensions['L'].width=58
    ws.column_dimensions['M'].width=8; ws.column_dimensions['N'].width=48
    ws.freeze_panes='B5'

    # ==== YTD PL ====
    ws2=wb.create_sheet('Consolidated YTD PL'); ws2.sheet_properties.tabColor='7030A0'
    ws2.merge_cells('A1:H1'); ws2.cell(1,1,'CONSOLIDATED YTD P&L — Variance Analysis').font=TF
    ws2.merge_cells('A2:H2'); ws2.cell(2,1,f'Six Months Ended June 30, 2025 vs June 30, 2024 | Material: >{mp*100:.0f}% or >{ma/1000:.0f}K | {ts}').font=NF
    r=4
    for i,h in enumerate(['Account','Jun 30, 2025\nYTD','Jun 30, 2024\nYTD','YTD $','YTD %','YTD Explanation','Cov%','Follow-Up Flag'],1):
        c=ws2.cell(r,i,h); c.font=WB; c.fill=DP; c.border=TB
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws2.row_dimensions[r].height=40; r+=1
    for a,d in cy:
        if a=='__S__':
            ws2.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
            ws2.cell(r,1,d).font=SF
            for cc in range(1,9): ws2.cell(r,cc).fill=DH; ws2.cell(r,cc).border=TB
            r+=1; continue
        ie=d['section'] in('Cost of Revenue','Operating Expenses')
        v,p=cv(d['ytd'],d['ytdy'],ie); m=im(v,p,mp,ma)
        ws2.cell(r,1,a).font=B; ws2.cell(r,1).border=TB
        ws2.cell(r,2,d['ytd']).font=B; ws2.cell(r,2).number_format=AF; ws2.cell(r,2).border=TB
        ws2.cell(r,3,d['ytdy']).font=B; ws2.cell(r,3).number_format=AF; ws2.cell(r,3).border=TB
        vc(ws2,r,4,v,m)
        c=ws2.cell(r,5,p if p else 0); c.number_format=PF; c.border=TB; c.font=B
        w_expl(ws2,r,6,d.get('ytd_expl',[]),gaf(cm,'ytd',a,'explanation'),m)
        tcov=gaf(cm,'ytd',a,'coverage_pct',None)
        w_cov(ws2,r,7,tcov,m)
        fu=gaf(cm,'ytd',a,'followup','') if m else ''
        w_fu(ws2,r,8,fu)
        if fu:
            dl=gaf(cm,'ytd',a,'followup_depts',[])
            all_fu.append({'tab':'YTD PL','period':'YTD','account':a,'issue':fu,
                'depts':dl or list(set(d['depts']))})
        r+=1
    ws2.column_dimensions['A'].width=35; ws2.column_dimensions['B'].width=16; ws2.column_dimensions['C'].width=16
    ws2.column_dimensions['D'].width=13; ws2.column_dimensions['E'].width=9; ws2.column_dimensions['F'].width=65
    ws2.column_dimensions['G'].width=8; ws2.column_dimensions['H'].width=48
    ws2.freeze_panes='B5'

    # ==== BS ====
    ws3=wb.create_sheet('Consolidated BS'); ws3.sheet_properties.tabColor='548235'
    ws3.merge_cells('A1:H1'); ws3.cell(1,1,'CONSOLIDATED BALANCE SHEET — Variance Analysis').font=TF
    ws3.merge_cells('A2:H2'); ws3.cell(2,1,f'As of June 30, 2025 vs December 31, 2024 | Material: >{mp*100:.0f}% or >{ma/1000:.0f}K | {ts}').font=NF
    r=4
    for i,h in enumerate(['Account','Jun 30, 2025','Dec 31, 2024','Change $','Change %','Explanation','Cov%','Follow-Up Flag'],1):
        c=ws3.cell(r,i,h); c.font=WB; c.fill=DG; c.border=TB
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws3.row_dimensions[r].height=40; r+=1
    for a,d in cb:
        if a=='__S__':
            ws3.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
            ws3.cell(r,1,d).font=SF
            for cc in range(1,9): ws3.cell(r,cc).fill=DH; ws3.cell(r,cc).border=TB
            r+=1; continue
        v=d['jun']-d['dec']; p=v/abs(d['dec']) if d['dec']!=0 else None; m=im(v,p,mp,ma)
        ws3.cell(r,1,a).font=B; ws3.cell(r,1).border=TB
        ws3.cell(r,2,d['jun']).font=B; ws3.cell(r,2).number_format=AF; ws3.cell(r,2).border=TB
        ws3.cell(r,3,d['dec']).font=B; ws3.cell(r,3).number_format=AF; ws3.cell(r,3).border=TB
        vc(ws3,r,4,v,m)
        c=ws3.cell(r,5,p if p else 0); c.number_format=PF; c.border=TB; c.font=B
        w_expl(ws3,r,6,d.get('expl',[]),gaf(cm,'bs',a,'explanation'),m)
        bcov=gaf(cm,'bs',a,'coverage_pct',None)
        w_cov(ws3,r,7,bcov,m)
        fu=gaf(cm,'bs',a,'followup','') if m else ''
        w_fu(ws3,r,8,fu)
        if fu:
            dl=gaf(cm,'bs',a,'followup_depts',[])
            all_fu.append({'tab':'BS','period':'BS','account':a,'issue':fu,
                'depts':dl or list(set(d['depts']))})
        r+=1
    ws3.column_dimensions['A'].width=35; ws3.column_dimensions['B'].width=18; ws3.column_dimensions['C'].width=18
    ws3.column_dimensions['D'].width=14; ws3.column_dimensions['E'].width=9; ws3.column_dimensions['F'].width=65
    ws3.column_dimensions['G'].width=8; ws3.column_dimensions['H'].width=48
    ws3.freeze_panes='B5'

    # ==== FOLLOW-UP ACTIONS ====
    ws4=wb.create_sheet('Follow-Up Actions'); ws4.sheet_properties.tabColor='FF6600'
    ws4.merge_cells('A1:G1'); ws4.cell(1,1,'FOLLOW-UP ACTIONS — Items Requiring Department Response').font=TF
    ws4.merge_cells('A2:G2'); ws4.cell(2,1,f'Generated {ts} | Grouped by department for email follow-up').font=NF
    r=4
    if all_fu:
        di={}
        for f in all_fu:
            for dept in f['depts']:
                if dept not in di: di[dept]=[]
                di[dept].append(f)
        for i,h in enumerate(['Department','Statement','Period','Account','Issue Identified','Action Required','Status'],1):
            c=ws4.cell(r,i,h); c.font=WB; c.fill=OR_H; c.border=TB
            c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws4.row_dimensions[r].height=35; r+=1
        for dept in sorted(di.keys()):
            ws4.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
            ws4.cell(r,1,dept).font=Font(name='Arial',color='FFFFFF',size=11,bold=True)
            for cc in range(1,8): ws4.cell(r,cc).fill=DH; ws4.cell(r,cc).border=TB
            r+=1
            for f in di[dept]:
                ws4.cell(r,1,dept).font=B; ws4.cell(r,1).border=TB
                ws4.cell(r,2,f['tab']).font=B; ws4.cell(r,2).border=TB
                ws4.cell(r,3,f['period']).font=B; ws4.cell(r,3).border=TB
                ws4.cell(r,4,f['account']).font=BB; ws4.cell(r,4).border=TB
                ws4.cell(r,5,f['issue']).font=OF; ws4.cell(r,5).border=TB; ws4.cell(r,5).alignment=Alignment(wrap_text=True)
                act='Please review and provide corrected explanation.'
                iu=f['issue'].upper()
                if 'DIRECTION' in iu: act='Explanation direction conflicts with actual data. Confirm correct direction and resubmit with corrected dollar amounts.'
                elif 'MATH' in iu: act='Dollar amounts do not reconcile to variance. Recheck driver amounts and resubmit.'
                elif 'CONTRADICT' in iu: act='Explanation contradicts the actual balance movement. Verify balances and resubmit.'
                elif 'COVERAGE' in iu or 'INSUFFICIENT' in iu: act='Explanation covers less than 80% of variance. Identify additional drivers with dollar amounts.'
                ws4.cell(r,6,act).font=B; ws4.cell(r,6).border=TB; ws4.cell(r,6).alignment=Alignment(wrap_text=True)
                ws4.cell(r,7,'OPEN').font=Font(name='Arial',color='FF0000',size=10,bold=True)
                ws4.cell(r,7).border=TB; ws4.cell(r,7).alignment=Alignment(horizontal='center')
                r+=1
            r+=1
    else:
        ws4.cell(r,1,'No follow-up items. All explanations passed validation.').font=Font(name='Arial',color='006100',size=11)
    ws4.column_dimensions['A'].width=22; ws4.column_dimensions['B'].width=10; ws4.column_dimensions['C'].width=8
    ws4.column_dimensions['D'].width=30; ws4.column_dimensions['E'].width=52; ws4.column_dimensions['F'].width=48; ws4.column_dimensions['G'].width=10

    wb.save(op); print(f"\nSaved: {op}\n  Follow-ups: {len(all_fu)}")

def main():
    pa=argparse.ArgumentParser(description='Consolidated Variance AI Agent v5',formatter_class=argparse.RawDescriptionHelpFormatter)
    pa.add_argument('input_file',help='Dept submissions Excel file')
    pa.add_argument('-o','--output',default='consolidated_variance_report.xlsx',help='Output file')
    pa.add_argument('--materiality-pct',type=float,default=MPCT,help='Materiality pct (default 0.05)')
    pa.add_argument('--materiality-abs',type=float,default=MABS,help='Materiality dollars (default 500000)')
    pa.add_argument('--api-key',default=None,help='Anthropic API key')
    pa.add_argument('--no-ai',action='store_true',help='Skip AI')
    a=pa.parse_args()
    if not os.path.exists(a.input_file): print(f"ERROR: {a.input_file} not found"); sys.exit(1)
    mp=a.materiality_pct; ma=a.materiality_abs
    print("="*60); print("CONSOLIDATED VARIANCE ANALYSIS - AI AGENT v5"); print("="*60)
    print(f"Input:  {a.input_file}\nOutput: {a.output}\nMateriality: >{mp*100:.0f}% OR >{ma/1000:.0f}K\n")
    print("Step 1: Parsing..."); ds=parse(a.input_file)
    for n,d in ds.items(): print(f"  {n} ({d['type']}): {len(d['qtr'])} qtr, {len(d['ytd'])} ytd, {len(d['bs'])} bs")
    print("\nStep 2: Consolidating...")
    cq=consol(ds,'qtr',['q2','q1','q2y'],['qoq_expl','yoy_expl'])
    cy=consol(ds,'ytd',['ytd','ytdy'],['ytd_expl'])
    cb=consol(ds,'bs',['jun','dec'],['expl'])
    print(f"  Qtr: {sum(1 for x,_ in cq if x!='__S__')} | YTD: {sum(1 for x,_ in cy if x!='__S__')} | BS: {sum(1 for x,_ in cb if x!='__S__')}")
    cm={}
    if a.no_ai: print("\nStep 3: Skipped (--no-ai)")
    elif not HAS_API: print("\nStep 3: Skipped (no anthropic)")
    else:
        print("\nStep 3: AI commentary...")
        ak=get_key(a.api_key)
        if ak: os.environ['ANTHROPIC_API_KEY']=ak; cm=get_ai(cq,cy,cb,mp,ma)
        else: print("  No key.")
    print("\nStep 4: Building output...")
    build(cq,cy,cb,cm,ds,a.output,mp,ma)
    print(f"\n{'='*60}\nDONE | {len(ds)} depts | {a.output}\n{'='*60}")

if __name__=='__main__': main()